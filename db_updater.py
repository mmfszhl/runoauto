import os
import sys
import mysql.connector
from mysql.connector import errorcode
import pandas as pd
from openpyxl import load_workbook
from deepdiff import DeepDiff


sys.path.insert(0, os.path.dirname(__file__))


def app(environ, start_response):
    start_response('200 OK', [('Content-Type', 'text/plain')])
    message = 'Successfully updated!\n'
    version = 'Python v' + sys.version.split()[0] + '\n'
    response = '\n'.join([message, version])
    return [response.encode()]


def read_excel(filename):
    '''Reading excel'''

    df = pd.read_excel(filename)
    df.to_excel("output.xlsx")

    wb = load_workbook(filename="output.xlsx")
    ws = wb.active
    goods_dict = {}
    for row_value in range(10, ws.max_row + 1):
        existance_cell = ws.cell(row=row_value, column=15).value
        if existance_cell is not None:
            if existance_cell == "1":
                stock_status = 'instock'
            else:
                stock_status = 'outofstock'

            #price = round(float(ws.cell(row=row_value, column=5).value), 3)
            woosale = round(float(ws.cell(row=row_value, column=6).value), 3)
            wcwp_sto = round(float(ws.cell(row=row_value, column=7).value), 3)
            regular_price = round(float(ws.cell(row=row_value, column=8).value), 3)
            tmp_dict = {'_stock_status': stock_status, 'wcwp_wholesale': woosale, 'wcwp_sto': wcwp_sto,
                        '_regular_price': regular_price}
            goods_dict.update({ws.cell(row=row_value, column=2).value: tmp_dict})

    wb.close()
    if os.path.exists("output.xlsx"):
        os.remove("output.xlsx")
    else:
        print("The file does not exist")

    #print(goods_dict)
    return goods_dict


def create_connection():
    '''Creating connection to local DB'''

    con = None
    try:
        con = mysql.connector.connect(
            user='runoavto_db',
            password='lNBmBVYBK$Ns',
            host='localhost',
            database='runoavto_db'
        )
        print("Successfully connected")
        return con
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)


def select_postid_with_sku(conn):
    '''Select post_id by _sku'''

    cur = conn.cursor()
    postid_dict = {}
    cur.execute("SELECT * FROM `wp_postmeta` WHERE `meta_key` = '_sku'")
    rows = cur.fetchall()
    for row in rows:
        postid_dict.update({row[3]: row[1]})

    return postid_dict


def select_data_with_postid(conn, postid_dict):
    '''Select prices by post_id'''

    cur = conn.cursor()
    data_dict = {}
    for article, post_id in postid_dict.items():
        cur.execute("SELECT * FROM `wp_postmeta` WHERE `post_id`=%s", (post_id, ))
        rows = cur.fetchall()
        stock, price, woosale, regular_price, wcwp_sto = 'outofstock', 0, 0, 0, 0
        for row in rows:
            meta_key = row[2]
            meta_value = row[3]
            if meta_key == "_stock_status":
                if meta_value == "instock" or meta_value == 1 or meta_value == "1":
                    stock = 'instock'
            elif meta_key == "_price":
                if meta_value is not None and meta_value != '':
                    price = round(float(meta_value), 3)
            elif meta_key == "wcwp_wholesale":
                if meta_value is not None and meta_value != '':
                    woosale = round(float(meta_value), 3)
            elif meta_key == "_regular_price":
                if meta_value is not None and meta_value != '':
                    regular_price = round(float(meta_value), 3)
            elif meta_key == "wcwp_sto":
                if meta_value is not None and meta_value != '':
                    wcwp_sto = round(float(meta_value), 3)

        tmp_dict = {'_stock_status': stock, 'wcwp_wholesale': woosale, 'wcwp_sto': wcwp_sto,
                    '_regular_price': regular_price}
        data_dict.update({article: tmp_dict})

    return data_dict


def select_metaid(conn, post_id, meta_key):
    '''Select meta_id by post_id and meta_key'''

    cur = conn.cursor()
    cur.execute("SELECT * FROM `wp_postmeta` WHERE `post_id` =%s and `meta_key` =%s", (post_id, meta_key))
    row = cur.fetchone()
    print('old_row', row)
    return row[0]


def check_update(conn, post_id, meta_key):
    '''Select meta_id by post_id and meta_key'''

    cur = conn.cursor()
    cur.execute("SELECT * FROM `wp_postmeta` WHERE `post_id` =%s and `meta_key` =%s", (post_id, meta_key))
    row = cur.fetchone()
    print('new_row', row)
    return row[0]


def update_row(conn, meta_value, meta_id):
    '''Update meta_value by meta_id'''
    cur = conn.cursor()
    cur.execute("UPDATE `wp_postmeta` SET `meta_value` =%s WHERE `wp_postmeta`.`meta_id` =%s", (meta_value, meta_id))
    conn.commit()

def create_ldiff(ddiff):
    '''Creating list of articles with differences'''
    ldiff = []
    for k, v in ddiff.items():
        if k == "values_changed":
            old_art = ""
            for i, j in v.items():
                txt = i.replace("]", "").replace("root[", "").replace("'", "")
                txt_lst = txt.split("[")
                new_art = txt_lst[0]
                if new_art != old_art:
                    ldiff.append(new_art)
                    old_art = new_art
    print("difflen", len(ldiff))
    return ldiff


def update_table(conn, ldiff, postid_dict, ex_dict):
    '''Update all rows with arcitle from ldiff'''

    for article in ldiff:
        post_id = postid_dict.get(article)
        #print(post_id)
        for meta_key, meta_value in ex_dict.get(article).items():
            #print('meta_key', meta_key, 'meta_value', meta_value)
            meta_id = select_metaid(conn, post_id, meta_key)
            #print('meta_id', meta_id)
            update_row(conn, meta_value, meta_id)
            check_update(conn, post_id, meta_key)


def main():
    # create a database connection
    conn = create_connection()
    ex_dict = read_excel(r'/home/runoavto/public_html/runoauto_price_new.xlsx')
    with conn:
        postid_dict = select_postid_with_sku(conn)
        db_dict = select_data_with_postid(conn, postid_dict)
        #print("db_dict", db_dict)
        #print("ex_dict", ex_dict)
        ddiff = DeepDiff(db_dict, ex_dict, ignore_order=True, ignore_string_case=True, ignore_numeric_type_changes=True)
        #print(ddiff)
        ldiff = create_ldiff(ddiff)
        update_table(conn, ldiff, postid_dict, ex_dict)

    conn.close()
    print("FINISHED!")


if __name__ == '__main__':
    main()
    app()